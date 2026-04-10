"""Shared styles, helpers, and column auto-sizing for all Excel exporters."""
from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
ALT_ROW_FILL  = PatternFill("solid", fgColor="D6E4F0")
DENIED_FILL   = PatternFill("solid", fgColor="FFCCCC")
PARTIAL_FILL  = PatternFill("solid", fgColor="FFF2CC")
OVER_FILL     = PatternFill("solid", fgColor="FF9999")
UNDER_FILL    = PatternFill("solid", fgColor="FFFFCC")
OK_FILL       = PatternFill("solid", fgColor="CCFFCC")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Calibri", size=9)
THIN_BORDER   = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)


def apply_header_row(ws, headers: list[str], row: int = 1) -> None:
    """Write and style a header row."""
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def style_data_row(ws, row: int, num_cols: int, fill: PatternFill | None = None) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font   = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if fill:
            cell.fill = fill
        elif row % 2 == 0:
            cell.fill = ALT_ROW_FILL


def auto_size_columns(ws, max_width: int = 50) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, max_width)


def freeze_header(ws) -> None:
    ws.freeze_panes = "A2"


def new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    return wb


def save_workbook(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
