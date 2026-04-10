"""835 ERA Excel exporter."""
from __future__ import annotations
from datetime import date
from openpyxl.styles import PatternFill
from .base_excel import (
    apply_header_row, style_data_row, auto_size_columns, freeze_header,
    new_workbook, save_workbook, DENIED_FILL, PARTIAL_FILL
)

CLP_STATUS = {
    "1": "Processed as Primary",
    "2": "Processed as Other",
    "3": "Processed as Secondary",
    "4": "Denied",
    "19": "Processed as Primary, Forwarded to Additional Payer(s)",
    "20": "Processed as Secondary, Forwarded to Additional Payer(s)",
    "22": "Reversal of Previous Payment",
}


def _fmt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, date):
        return val.strftime("%m/%d/%Y")
    return str(val)


def export_835(parsed_data: dict) -> bytes:
    wb = new_workbook()
    header = parsed_data.get("header", {})
    claims = parsed_data.get("claim_payments", [])
    plb    = parsed_data.get("provider_adjustments", [])

    # ── Summary ─────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    apply_header_row(ws, ["Metric", "Value"])
    total_billed = sum(c.get("billed") or 0 for c in claims)
    total_paid   = sum(c.get("paid")   or 0 for c in claims)
    total_pr     = sum(c.get("patient_responsibility") or 0 for c in claims)
    denied_count = sum(1 for c in claims if c.get("status_code") == "4")
    rows = [
        ("Payer Name",           header.get("payer_name", "")),
        ("Payee NPI",            header.get("payee_npi", "")),
        ("Check / EFT Number",   header.get("check_eft_number", "")),
        ("Payment Date",         _fmt(header.get("payment_date"))),
        ("Total Payment ($)",    f"{header.get('total_payment') or 0:,.2f}"),
        ("Payment Method",       header.get("payment_method", "")),
        ("Total Claims",         len(claims)),
        ("Denied Claims",        denied_count),
        ("Denial Rate",          f"{denied_count/len(claims)*100:.1f}%" if claims else "0%"),
        ("Total Billed ($)",     f"{total_billed:,.2f}"),
        ("Total Paid ($)",       f"{total_paid:,.2f}"),
        ("Patient Responsibility ($)", f"{total_pr:,.2f}"),
    ]
    for r, (metric, val) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=metric)
        ws.cell(row=r, column=2, value=val)
        style_data_row(ws, r, 2)
    auto_size_columns(ws); freeze_header(ws)

    # ── Claim Payments ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Claim Payments")
    hdrs = ["CLP ID", "Status Code", "Status", "Patient Name", "Patient ID",
            "Billed ($)", "Paid ($)", "Patient Resp. ($)", "Claim Filing Indicator",
            "Payer Claim #"]
    apply_header_row(ws, hdrs)
    for r, c in enumerate(claims, start=2):
        status_code = c.get("status_code", "")
        row = [
            c.get("clp_id", ""), status_code,
            CLP_STATUS.get(status_code, status_code),
            c.get("patient_name", ""), c.get("patient_id", ""),
            c.get("billed", ""), c.get("paid", ""),
            c.get("patient_responsibility", ""),
            c.get("claim_filing_indicator", ""),
            c.get("payer_claim_number", ""),
        ]
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)
        fill = DENIED_FILL if status_code == "4" else (PARTIAL_FILL if status_code in ("2","3") else None)
        style_data_row(ws, r, len(hdrs), fill=fill)
    auto_size_columns(ws); freeze_header(ws)

    # ── Service Line Payments ────────────────────────────────────────────────────
    ws = wb.create_sheet("Service Line Payments")
    hdrs = ["CLP ID", "CPT/HCPCS", "Billed ($)", "Paid ($)", "NDC",
            "CO Adjustments ($)", "PR Adjustments ($)", "OA Adjustments ($)"]
    apply_header_row(ws, hdrs)
    row_num = 2
    for c in claims:
        for svc in c.get("services", []):
            adjs = svc.get("adjustments", [])
            co = sum(a.get("amount") or 0 for a in adjs if a.get("group_code") == "CO")
            pr = sum(a.get("amount") or 0 for a in adjs if a.get("group_code") == "PR")
            oa = sum(a.get("amount") or 0 for a in adjs if a.get("group_code") == "OA")
            row = [c.get("clp_id",""), svc.get("cpt_hcpcs",""),
                   svc.get("billed",""), svc.get("paid",""), svc.get("ndc",""),
                   co or "", pr or "", oa or ""]
            for col, val in enumerate(row, start=1):
                ws.cell(row=row_num, column=col, value=val)
            style_data_row(ws, row_num, len(hdrs))
            row_num += 1
    auto_size_columns(ws); freeze_header(ws)

    # ── Adjustments ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Adjustments")
    hdrs = ["CLP ID", "Level", "Group Code", "Group Description",
            "Reason Code", "Amount ($)"]
    apply_header_row(ws, hdrs)
    row_num = 2
    for c in claims:
        for adj in c.get("adjustments", []):
            row = [c.get("clp_id",""), "Claim",
                   adj.get("group_code",""), adj.get("group_description",""),
                   adj.get("reason_code",""), adj.get("amount","")]
            for col, val in enumerate(row, start=1):
                ws.cell(row=row_num, column=col, value=val)
            style_data_row(ws, row_num, len(hdrs))
            row_num += 1
        for svc in c.get("services", []):
            for adj in svc.get("adjustments", []):
                row = [c.get("clp_id",""), f"SVC {svc.get('cpt_hcpcs','')}",
                       adj.get("group_code",""), adj.get("group_description",""),
                       adj.get("reason_code",""), adj.get("amount","")]
                for col, val in enumerate(row, start=1):
                    ws.cell(row=row_num, column=col, value=val)
                style_data_row(ws, row_num, len(hdrs))
                row_num += 1
    auto_size_columns(ws); freeze_header(ws)

    # ── Denial Analysis ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Denial Analysis")
    hdrs = ["Reason Code", "Group Code", "Count", "Total Amount ($)", "% of All Denials"]
    apply_header_row(ws, hdrs)
    from collections import defaultdict
    denial_map: dict[tuple, dict] = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for c in claims:
        for adj in c.get("adjustments", []):
            key = (adj.get("reason_code",""), adj.get("group_code",""))
            denial_map[key]["count"]  += 1
            denial_map[key]["amount"] += adj.get("amount") or 0
    total_denials = sum(v["count"] for v in denial_map.values())
    for r, ((rc, gc), vals) in enumerate(
        sorted(denial_map.items(), key=lambda x: -x[1]["count"]), start=2
    ):
        pct = f"{vals['count']/total_denials*100:.1f}%" if total_denials else "0%"
        row = [rc, gc, vals["count"], f"{vals['amount']:,.2f}", pct]
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)
        style_data_row(ws, r, len(hdrs))
    auto_size_columns(ws); freeze_header(ws)

    # ── Provider Adjustments (PLB) ───────────────────────────────────────────────
    if plb:
        ws = wb.create_sheet("Provider Adjustments")
        hdrs = ["Provider NPI", "Fiscal Period", "Reason Code", "Reference ID", "Amount ($)"]
        apply_header_row(ws, hdrs)
        for r, p in enumerate(plb, start=2):
            row = [p.get("provider_npi",""), p.get("fiscal_period",""),
                   p.get("reason_code",""), p.get("reference_id",""), p.get("amount","")]
            for col, val in enumerate(row, start=1):
                ws.cell(row=r, column=col, value=val)
            style_data_row(ws, r, len(hdrs))
        auto_size_columns(ws); freeze_header(ws)

    # ── Reconciliation ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Reconciliation")
    hdrs = ["CLP ID", "Billed ($)", "Paid ($)", "Total Adjustments ($)",
            "Calculated Total ($)", "Match?"]
    apply_header_row(ws, hdrs)
    for r, c in enumerate(claims, start=2):
        billed = c.get("billed") or 0
        paid   = c.get("paid") or 0
        all_adjs: list[dict] = list(c.get("adjustments", []))
        for svc in c.get("services", []):
            all_adjs.extend(svc.get("adjustments", []))
        total_adj = sum(a.get("amount") or 0 for a in all_adjs)
        calc = round(paid + total_adj, 2)
        match = "YES" if abs(calc - billed) < 0.02 else "NO"
        fill = DENIED_FILL if match == "NO" else None
        row = [c.get("clp_id",""), billed, paid, round(total_adj,2), calc, match]
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)
        style_data_row(ws, r, len(hdrs), fill=fill)
    auto_size_columns(ws); freeze_header(ws)

    return save_workbook(wb)
