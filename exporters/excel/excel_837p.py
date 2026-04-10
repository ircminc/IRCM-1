"""837P Excel exporter — multi-tab workbook."""
from __future__ import annotations
from datetime import date
from openpyxl import Workbook
from .base_excel import (
    apply_header_row, style_data_row, auto_size_columns, freeze_header,
    new_workbook, save_workbook, DENIED_FILL, OVER_FILL, UNDER_FILL, OK_FILL
)

CMS_HEADERS = [
    "CPT/HCPCS", "Modifier", "Description", "Billed Amount",
    "Medicare Non-Facility Rate", "Medicare Facility Rate", "Work RVU",
    "ASP+6% (Drug)", "Vs Non-Fac %", "Vs Facility %", "Rate Flag", "Rate Source"
]

CLAIM_HEADERS = [
    "Claim ID", "Patient Last", "Patient First", "Patient DOB", "Gender",
    "Subscriber ID", "Group Number", "Billing Provider NPI", "Billing Provider Name",
    "Payer ID", "Payer Name", "DOS From", "DOS To", "Place of Service",
    "Claim Frequency", "Total Billed ($)", "Claim Filing Indicator",
    "Principal Dx", "Secondary Dx 1", "Secondary Dx 2", "Secondary Dx 3",
    "Claim Note", "Payer Claim #"
]

SVC_HEADERS = [
    "Claim ID", "Line #", "CPT/HCPCS", "Mod 1", "Mod 2", "Mod 3", "Mod 4",
    "Billed Amount ($)", "Units", "Place of Service", "Dx Pointers",
    "NDC", "Rendering Provider NPI", "Rendering Provider"
]

PROVIDER_HEADERS = [
    "Type", "NPI", "Name", "Entity Type", "Tax ID", "Address", "Phone", "Taxonomy"
]

DX_HEADERS = [
    "Claim ID", "Code Set", "Principal Dx",
    "Dx 2", "Dx 3", "Dx 4", "Dx 5", "Dx 6",
    "Dx 7", "Dx 8", "Dx 9", "Dx 10", "Dx 11", "Dx 12"
]


def _fmt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, date):
        return val.strftime("%m/%d/%Y")
    return str(val)


def export_837p(parsed_data: dict, cms_comparisons: list[dict] | None = None) -> bytes:
    """
    parsed_data: output of parse_837p()  i.e. {"claims": [...], "providers": [...]}
    cms_comparisons: optional list of RateComparison dicts keyed by (claim_id, line_number)
    Returns raw .xlsx bytes.
    """
    wb = new_workbook()
    claims = parsed_data.get("claims", [])
    providers = parsed_data.get("providers", [])

    # ── Tab 1: Summary ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    apply_header_row(ws, ["Metric", "Value"])
    summary_rows = [
        ("Total Claims", len(claims)),
        ("Total Billed ($)", f"{sum(c.get('total_billed') or 0 for c in claims):,.2f}"),
        ("Unique Billing Providers", len({c.get('billing_provider', {}).get('npi') for c in claims} - {''})),
        ("Unique Payers", len({c.get('payer_id') for c in claims} - {''})),
        ("Service Lines", sum(len(c.get('service_lines', [])) for c in claims)),
        ("DOS Range From", _fmt(min((c['dos_from'] for c in claims if c.get('dos_from')), default=None))),
        ("DOS Range To",   _fmt(max((c['dos_to'] or c.get('dos_from') for c in claims if c.get('dos_from')), default=None))),
    ]
    for r, (metric, val) in enumerate(summary_rows, start=2):
        ws.cell(row=r, column=1, value=metric)
        ws.cell(row=r, column=2, value=val)
        style_data_row(ws, r, 2)
    auto_size_columns(ws); freeze_header(ws)

    # ── Tab 2: Claims ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Claims")
    apply_header_row(ws, CLAIM_HEADERS)
    for r, c in enumerate(claims, start=2):
        dxs = c.get("diagnoses", [])
        row = [
            c.get("claim_id", ""),
            c.get("patient", {}).get("last_name", "") or c.get("subscriber", {}).get("last_name", ""),
            c.get("patient", {}).get("first_name", "") or c.get("subscriber", {}).get("first_name", ""),
            _fmt(c.get("patient", {}).get("dob") or c.get("subscriber", {}).get("dob")),
            c.get("patient", {}).get("gender", "") or c.get("subscriber", {}).get("gender", ""),
            c.get("subscriber", {}).get("member_id", ""),
            c.get("group_number", ""),
            c.get("billing_provider", {}).get("npi", ""),
            c.get("billing_provider", {}).get("last_name_org", ""),
            c.get("payer_id", ""), c.get("payer_name", ""),
            _fmt(c.get("dos_from")), _fmt(c.get("dos_to")),
            c.get("place_of_service", ""), c.get("claim_frequency", ""),
            c.get("total_billed", ""),
            c.get("claim_filing_indicator", ""),
            dxs[0]["code"] if len(dxs) > 0 else "",
            dxs[1]["code"] if len(dxs) > 1 else "",
            dxs[2]["code"] if len(dxs) > 2 else "",
            dxs[3]["code"] if len(dxs) > 3 else "",
            c.get("claim_note", ""),
            c.get("payer_claim_number", ""),
        ]
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)
        style_data_row(ws, r, len(CLAIM_HEADERS))
    auto_size_columns(ws); freeze_header(ws)

    # ── Tab 3: Service Lines ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Service Lines")
    apply_header_row(ws, SVC_HEADERS)
    row_num = 2
    for c in claims:
        for sl in c.get("service_lines", []):
            row = [
                c.get("claim_id", ""),
                sl.get("line_number", ""),
                sl.get("cpt_hcpcs", ""),
                sl.get("modifier_1", ""), sl.get("modifier_2", ""),
                sl.get("modifier_3", ""), sl.get("modifier_4", ""),
                sl.get("billed_amount", ""),
                sl.get("units", ""),
                sl.get("place_of_service", ""),
                sl.get("diagnosis_pointers", ""),
                sl.get("ndc", ""),
                sl.get("rendering_provider_npi", ""),
                sl.get("rendering_provider_name", ""),
            ]
            for col, val in enumerate(row, start=1):
                ws.cell(row=row_num, column=col, value=val)
            style_data_row(ws, row_num, len(SVC_HEADERS))
            row_num += 1
    auto_size_columns(ws); freeze_header(ws)

    # ── Tab 4: Providers ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Providers")
    apply_header_row(ws, PROVIDER_HEADERS)
    for r, p in enumerate(providers, start=2):
        row = [
            p.get("type", ""), p.get("npi", ""), p.get("name", ""),
            p.get("entity_type", ""), p.get("tax_id", ""), p.get("address", ""),
            p.get("phone", ""), p.get("taxonomy", ""),
        ]
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)
        style_data_row(ws, r, len(PROVIDER_HEADERS))
    auto_size_columns(ws); freeze_header(ws)

    # ── Tab 5: Diagnoses ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Diagnoses")
    apply_header_row(ws, DX_HEADERS)
    for r, c in enumerate(claims, start=2):
        dxs = c.get("diagnoses", [])
        codes = [d["code"] for d in dxs]
        qualifier = dxs[0]["qualifier"] if dxs else ""
        row = [c.get("claim_id", ""), qualifier] + (codes + [""] * 12)[:12]
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)
        style_data_row(ws, r, len(DX_HEADERS))
    auto_size_columns(ws); freeze_header(ws)

    # ── Tab 6: CMS Comparison ────────────────────────────────────────────────────
    if cms_comparisons:
        ws = wb.create_sheet("CMS Comparison")
        apply_header_row(ws, CMS_HEADERS)
        for r, comp in enumerate(cms_comparisons, start=2):
            flag = comp.get("flag", "")
            row = [
                comp.get("cpt_hcpcs", ""),
                comp.get("modifier", ""),
                comp.get("description", ""),
                comp.get("billed_amount", ""),
                comp.get("pfs_non_facility_rate", ""),
                comp.get("pfs_facility_rate", ""),
                comp.get("work_rvu", ""),
                comp.get("asp_payment_limit", ""),
                f"{comp.get('vs_non_facility_pct', ''):.1f}%" if comp.get("vs_non_facility_pct") is not None else "",
                f"{comp.get('vs_facility_pct', ''):.1f}%"    if comp.get("vs_facility_pct") is not None else "",
                flag,
                comp.get("rate_source", ""),
            ]
            for col, val in enumerate(row, start=1):
                ws.cell(row=r, column=col, value=val)
            fill = OVER_FILL if flag == "OVER_300PCT" else (UNDER_FILL if flag == "UNDER_100PCT" else OK_FILL)
            style_data_row(ws, r, len(CMS_HEADERS), fill=fill)
        auto_size_columns(ws); freeze_header(ws)

    return save_workbook(wb)
